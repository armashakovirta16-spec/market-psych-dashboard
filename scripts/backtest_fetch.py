"""
Fetches and caches full free historical series for the item-4 backtest —
kept separate from fetch_data.py (the live daily pipeline) since these are
one-off-ish bulk historical pulls, not something to re-run on every Actions
run. Each series is cached to data/backtest_raw/*.csv so re-running
backtest.py doesn't re-hit these sources every time.

See memory/backtest_data_availability.md (session memory) for the verified
free-history depth of each source — two real gaps found vs. Will's memo
assumption: ISM PMI has no free full history anywhere, and HY OAS's free
FRED history only goes back to ~2023 (ICE licensing).
"""

import csv
import io
import os
import re

import openpyxl
import pandas as pd
import requests
import xlrd
import yfinance as yf

RAW_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "backtest_raw")
os.makedirs(RAW_DIR, exist_ok=True)

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}


def _cache_path(name):
    return os.path.join(RAW_DIR, f"{name}.csv")


def _write_series(name, dates, values, value_col="value"):
    path = _cache_path(name)
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["date", value_col])
        for d, v in zip(dates, values):
            writer.writerow([d, v])
    print(f"[{name}] wrote {len(dates)} rows -> {path}")


def fetch_yfinance_history(ticker, name):
    hist = yf.Ticker(ticker).history(period="max")
    closes = hist["Close"].dropna()
    dates = [d.strftime("%Y-%m-%d") for d in closes.index]
    _write_series(name, dates, closes.values, "close")


def fetch_fred_series(series_id, name):
    """Free CSV export, no API key — see fetch_data.py's fetch_consumer_sentiment
    docstring for why this route works for any FRED series, not just UMCSENT."""
    url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
    df = pd.read_csv(url, index_col=0, parse_dates=True)
    series = df[series_id].dropna()
    dates = [d.strftime("%Y-%m-%d") for d in series.index]
    _write_series(name, dates, series.values, "value")


def fetch_sp500_pe_monthly():
    """Full monthly S&P 500 P/E history from multpl.com back to 1871 —
    a different page than fetch_data.py's live scraper (which only reads
    today's value from the single-value page)."""
    resp = requests.get(
        "https://www.multpl.com/s-p-500-pe-ratio/table/by-month",
        headers=REQUEST_HEADERS,
        timeout=20,
    )
    resp.raise_for_status()
    # Rows look like <td>Jul 24, 2026</td>\n<td>\n<abbr title="Estimate">†</abbr>\n28.52\n</td>
    # (recent rows) or <td>Jan 1, 1900</td>\n<td>\n&#x2002;\n12.71\n</td> (older
    # rows use an HTML entity spacer instead of the <abbr> marker) — capture the
    # whole second cell loosely, then pull the trailing number out of it.
    rows = re.findall(
        r"<td>([A-Za-z]+ \d{1,2}, \d{4})</td>\s*<td>(.*?)</td>",
        resp.text,
        re.S,
    )
    if not rows:
        raise ValueError("No P/E rows found — multpl.com table structure may have changed")
    dates, values = [], []
    for date_str, cell in rows:
        value_match = re.search(r"([\d.]+)\s*$", cell.strip())
        if not value_match:
            continue
        dt = pd.to_datetime(date_str)
        dates.append(dt.strftime("%Y-%m-%d"))
        values.append(float(value_match.group(1)))
    # Page lists newest-first; store chronologically.
    dates, values = dates[::-1], values[::-1]
    _write_series("sp500_pe_monthly", dates, values, "pe")


def fetch_put_call_historical():
    """Cboe's own downloadable archive + recent CSVs, combined — a
    different source than fetch_data.py's live scraper (which reads the
    current single value from the daily market-statistics HTML page)."""
    archive = requests.get(
        "https://cdn.cboe.com/resources/options/volume_and_call_put_ratios/totalpcarchive.csv",
        headers=REQUEST_HEADERS,
        timeout=20,
    )
    archive.raise_for_status()
    recent = requests.get(
        "https://cdn.cboe.com/resources/options/volume_and_call_put_ratios/totalpc.csv",
        headers=REQUEST_HEADERS,
        timeout=20,
    )
    recent.raise_for_status()

    records = {}
    for text in (archive.text, recent.text):
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            if len(row) < 5:
                continue
            date_str = row[0].strip()
            pc_str = row[4].strip()
            try:
                dt = pd.to_datetime(date_str)
            except (ValueError, TypeError):
                continue
            try:
                pc = float(pc_str)
            except ValueError:
                continue
            records[dt.strftime("%Y-%m-%d")] = pc

    dates = sorted(records)
    values = [records[d] for d in dates]
    _write_series("put_call_historical", dates, values, "total_pc")


def fetch_aaii_historical():
    """Full weekly AAII history back to ~1987 from the archive .xls —
    a different, second free source than fetch_data.py's live scraper
    (which reads only the current week from the survey results page)."""
    resp = requests.get(
        "https://www.aaii.com/files/surveys/sentiment.xls",
        headers=REQUEST_HEADERS,
        timeout=30,
    )
    resp.raise_for_status()
    wb = xlrd.open_workbook(file_contents=resp.content)
    sh = wb.sheet_by_name("SENTIMENT")

    path = os.path.join(RAW_DIR, "aaii_historical.csv")
    written = 0
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["date", "bullish", "neutral", "bearish"])
        for r in range(5, sh.nrows):
            row = sh.row_values(r)
            date_val, bullish, neutral, bearish = row[0], row[1], row[2], row[3]
            if not isinstance(date_val, (int, float)) or date_val <= 0:
                continue
            if not all(isinstance(v, (int, float)) for v in (bullish, neutral, bearish)):
                continue
            dt = xlrd.xldate.xldate_as_datetime(date_val, wb.datemode)
            writer.writerow([dt.strftime("%Y-%m-%d"), bullish, neutral, bearish])
            written += 1
    print(f"[aaii_historical] wrote {written} rows -> {path}")


def fetch_news_sentiment_historical():
    resp = requests.get(
        "https://www.frbsf.org/wp-content/uploads/news-sentiment-chart-1.csv",
        headers=REQUEST_HEADERS,
        timeout=20,
    )
    resp.raise_for_status()
    rows = list(csv.reader(io.StringIO(resp.text)))
    dates, values = [], []
    for row in rows[1:]:
        if len(row) < 2 or row[1] in ("", "."):
            continue
        dt = pd.to_datetime(row[0])
        dates.append(dt.strftime("%Y-%m-%d"))
        values.append(float(row[1]))
    _write_series("news_sentiment_historical", dates, values, "value")


def fetch_naaim_historical():
    """Full weekly NAAIM history since inception (~2006) from the same
    xlsx fetch_data.py's live fetch_naaim_exposure() reads — that function
    only keeps the latest row; this keeps all of them."""
    page = requests.get(
        "https://naaim.org/programs/naaim-exposure-index/",
        headers=REQUEST_HEADERS,
        timeout=20,
    )
    page.raise_for_status()
    match = re.search(r'href="(https://naaim\.org/wp-content/uploads/\d{4}/\d{2}/[^"]+\.xlsx)"', page.text)
    if not match:
        raise ValueError("NAAIM data file link not found — page structure may have changed")

    xlsx_resp = requests.get(match.group(1), headers=REQUEST_HEADERS, timeout=20)
    xlsx_resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_resp.content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    naaim_idx = header.index("NAAIM Number")

    dates, values = [], []
    for row in rows[1:]:
        if row[naaim_idx] is None or row[0] is None:
            continue
        dates.append(row[0].strftime("%Y-%m-%d"))
        values.append(float(row[naaim_idx]))
    dates, values = dates[::-1], values[::-1]
    _write_series("naaim_historical", dates, values, "naaim_number")


def main():
    print("Fetching yfinance histories...")
    fetch_yfinance_history("^VIX", "vix_historical")
    for ticker in ["SPY", "AGG", "GLD", "XLK", "XLF", "XLE", "XLV", "XLU"]:
        fetch_yfinance_history(ticker, f"etf_{ticker}")

    print("Fetching FRED series...")
    for series_id, name in [
        ("CPIAUCSL", "fred_cpi"),
        ("UNRATE", "fred_unemployment"),
        ("FEDFUNDS", "fred_fedfunds"),
        ("DGS10", "fred_dgs10"),
        ("DGS2", "fred_dgs2"),
        ("UMCSENT", "fred_consumer_sentiment"),
        ("BAMLH0A0HYM2", "fred_hy_oas"),
    ]:
        fetch_fred_series(series_id, name)

    print("Fetching S&P 500 P/E monthly history...")
    fetch_sp500_pe_monthly()

    print("Fetching Cboe put/call historical...")
    fetch_put_call_historical()

    print("Fetching AAII historical...")
    fetch_aaii_historical()

    print("Fetching SF Fed news sentiment historical...")
    fetch_news_sentiment_historical()

    print("Fetching NAAIM historical...")
    fetch_naaim_historical()

    print("Done.")


if __name__ == "__main__":
    main()
